from __future__ import annotations

from datetime import date

from openpyxl import load_workbook
from PySide6.QtGui import QPageSize
from PySide6.QtPdf import QPdfDocument
from PySide6.QtWidgets import QApplication

from event_checklist.backup import create_backup, create_manual_backup, create_rotating_auto_backup, restore_backup
from event_checklist.export import default_excel_filename, export_excel, next_available_excel_path
from event_checklist.pdf_export import (
    PdfOptions, _checklist_standard_columns, default_pdf_filename, export_calendar_pdf, export_checklist_pdf,
    export_settlement_pdf, next_available_pdf_path, settlement_header_summary,
)
from event_checklist.services import EventService


def _event(db):
    service = EventService(db)
    ids = [row["id"] for row in db.query("SELECT id FROM master_items ORDER BY id LIMIT 2")]
    return service.create_event("내보내기 행사", date(2026, 9, 1), None, ids)


def test_backup_and_restore(db, tmp_path):
    _event(db)
    backup = create_backup(db, tmp_path / "backup.db")
    db.execute("DELETE FROM events")
    assert db.one("SELECT COUNT(*) count FROM events")["count"] == 0
    restore_backup(db, backup)
    assert db.one("SELECT COUNT(*) count FROM events")["count"] == 1
    assert db.one("SELECT COUNT(*) count FROM event_tasks")["count"] == 2


def test_excel_export_is_single_styled_report_with_filters_and_print_options(db, tmp_path):
    event_id = _event(db)
    selected = db.one(
        "SELECT id FROM event_tasks WHERE event_id=? ORDER BY id LIMIT 1", (event_id,),
    )
    db.execute(
        "UPDATE event_tasks SET major='시스템',minor='음향',detail='현장 확인 세부내용' WHERE id=?",
        (selected["id"],),
    )
    xlsx = export_excel(
        db, tmp_path / "output.xlsx", event_id, "checklist",
        PdfOptions("A3", "LANDSCAPE"), "시스템", "음향",
    )
    workbook = load_workbook(xlsx, read_only=False, data_only=False)
    assert workbook.sheetnames == ["체크리스트"]
    sheet = workbook["체크리스트"]
    assert sheet["A1"].fill.fgColor.rgb.endswith("F25B24")
    assert sheet["A4"].value.endswith("시스템 > 음향")
    assert sheet["A7"].value == 1
    assert sheet["B7"].value == "시스템"
    assert sheet["C7"].value == "음향"
    assert sheet["E7"].value == "현장 확인 세부내용"
    assert sheet["E7"].alignment.horizontal == "left"
    assert sheet["D7"].alignment.horizontal == "center"
    assert str(sheet.page_setup.paperSize) == sheet.PAPERSIZE_A3
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.print_title_rows == "$1:$6"
    assert sheet.sheet_view.showGridLines is False
    workbook.close()


def test_a4_portrait_excel_expands_two_line_cells_for_large_checklist(db, tmp_path):
    event_id = _event(db)
    for index in range(30):
        db.execute(
            """INSERT INTO event_tasks(event_id,major,minor,name,detail,status,planned_start,due_date,sort_order)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (
                event_id, "시스템", "무대", f"추가 항목 {index + 1}",
                "길이가 긴 세부내용은 자동으로 줄 수를 계산해 글자가 잘리지 않도록 행 높이를 확장합니다. "
                "인쇄 용지 폭을 넘는 추가 설명도 마지막 글자까지 모두 표시되어야 합니다."
                if index == 0 else "현장 확인",
                "미착수", "2026-09-06", "2026-10-01", 1000 + index,
            ),
        )
    path = export_excel(
        db, tmp_path / "checklist-a4-portrait.xlsx", event_id, "checklist",
        PdfOptions("A4", "PORTRAIT"),
    )
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheet = workbook["체크리스트"]
    assert sheet["D7"].font.sz == 10
    assert "\n" in sheet["I7"].value  # I열(9열)=시작일\n업체 (수량·단위 두 열 추가로 이동)
    assert sheet["I7"].alignment.wrap_text is True
    assert min(sheet.row_dimensions[row].height for row in range(7, sheet.max_row + 1)) >= 34
    assert max(sheet.row_dimensions[row].height for row in range(7, sheet.max_row + 1)) > 34
    workbook.close()


def test_settlement_excel_has_pdf_like_layout_formulas_and_all_page_options(db, tmp_path):
    event_id = _event(db)
    db.execute(
        "UPDATE events SET budget=500039000,budget_tax_mode='INCLUDED' WHERE id=?", (event_id,),
    )
    task_id = db.one(
        "SELECT id FROM event_tasks WHERE event_id=? ORDER BY id LIMIT 1", (event_id,),
    )["id"]
    db.execute(
        "UPDATE event_tasks SET detail=?,quantity=2,unit='식',unit_price=125000,vat_type='TAXABLE' WHERE id=?",
        ("항목 오른쪽에 표시되는 세부내용", task_id),
    )
    for paper in ("A4", "A3"):
        for orientation in ("PORTRAIT", "LANDSCAPE"):
            path = export_excel(
                db, tmp_path / f"settlement-{paper}-{orientation}.xlsx", event_id, "settlement",
                PdfOptions(paper, orientation),
            )
            workbook = load_workbook(path, read_only=False, data_only=False)
            assert workbook.sheetnames == ["정산내역"]
            sheet = workbook["정산내역"]
            assert "전체 프로젝트 금액(VAT 포함)  500,039,000원" in sheet["J3"].value
            assert "정산 합계  275,000원 · 499,764,000원 남음" in sheet["J3"].value
            assert sheet["J3"].alignment.wrap_text is True
            assert str(sheet.page_setup.paperSize) == (
                sheet.PAPERSIZE_A4 if paper == "A4" else sheet.PAPERSIZE_A3
            )
            assert sheet.page_setup.orientation == orientation.lower()
            detail_column = 4 if paper == "A4" and orientation == "PORTRAIT" else 12
            assert sheet.cell(7, detail_column).value == "항목 오른쪽에 표시되는 세부내용"
            assert sheet.cell(7, detail_column).alignment.horizontal == "left"
            supply_column = 8 if detail_column == 4 else 7
            assert str(sheet.cell(7, supply_column).value).startswith("=ROUND(")
            assert any(
                str(sheet.cell(row, 1).value) == "전체 합계"
                for row in range(7, sheet.max_row + 1)
            )
            workbook.close()


def test_settlement_output_header_names_budget_comparison_and_balance(db):
    event_id = _event(db)
    task_id = db.one(
        "SELECT id FROM event_tasks WHERE event_id=? ORDER BY id LIMIT 1", (event_id,),
    )["id"]
    db.execute(
        "UPDATE event_tasks SET quantity=2,unit_price=100000,vat_type='TAXABLE' WHERE id=?", (task_id,),
    )
    db.execute(
        "UPDATE events SET budget=300000,budget_tax_mode='INCLUDED' WHERE id=?", (event_id,),
    )
    included = EventService(db).settlement_summary(event_id)
    assert settlement_header_summary(included) == (
        "전체 프로젝트 금액(VAT 포함)  300,000원\n정산 합계  220,000원 · 80,000원 남음"
    )
    db.execute(
        "UPDATE events SET budget=150000,budget_tax_mode='EXCLUDED' WHERE id=?", (event_id,),
    )
    excluded = EventService(db).settlement_summary(event_id)
    assert settlement_header_summary(excluded) == (
        "전체 프로젝트 금액(VAT 별도)  150,000원\n정산 공급가  200,000원 · 50,000원 부족"
    )
    db.execute("UPDATE events SET budget_tax_mode='UNSET' WHERE id=?", (event_id,))
    unset = EventService(db).settlement_summary(event_id)
    assert settlement_header_summary(unset).endswith("VAT 기준 선택 필요")


def test_excel_default_filename_and_same_day_sequence(db, tmp_path):
    event_id = _event(db)
    event = db.one("SELECT * FROM events WHERE id=?", (event_id,))
    name = default_excel_filename(
        event, "checklist", PdfOptions("A4", "PORTRAIT"), date(2026, 8, 13),
        major="시스템", minor="음향",
    )
    assert name.endswith("_시스템_음향_20260813_A4세로.xlsx")
    first = tmp_path / name
    first.write_bytes(b"first")
    second = next_available_excel_path(first)
    second.write_bytes(b"second")
    assert second.name.endswith("_2.xlsx")
    assert next_available_excel_path(first).name.endswith("_3.xlsx")


def test_auto_backup_rotation_keeps_ten_and_preserves_manual(db, tmp_path):
    backup_directory = tmp_path / "backups"
    manual = create_manual_backup(db, backup_directory)
    for index in range(12):
        db.execute("UPDATE contacts SET phone=? WHERE id=1", (str(index),))
        create_rotating_auto_backup(db, backup_directory, keep=10)
    assert manual.exists()
    assert len(list(backup_directory.glob("auto_event_flow_*.db"))) == 10


def test_checklist_and_settlement_pdf_export_all_paper_orientations(db, tmp_path):
    app = QApplication.instance() or QApplication([])
    event_id = _event(db)
    db.execute(
        "UPDATE events SET name=? WHERE id=?",
        ("2026 대한민국 지역상생 문화예술축제 개막식 및 시민참여 프로그램 통합운영 행사", event_id),
    )
    tasks = db.query("SELECT id FROM event_tasks WHERE event_id=? ORDER BY id", (event_id,))
    db.execute(
        "UPDATE event_tasks SET detail=?,quantity=2,unit='식',unit_price=125000,vat_type='TAXABLE' WHERE id=?",
        ("항목 바로 오른쪽에 표시되는 긴 세부내용 검증 문장입니다. 필요한 행만 두 줄로 확장됩니다.", tasks[0]["id"]),
    )
    expected_sizes = {
        ("A4", "PORTRAIT"): QPageSize.PageSizeId.A4,
        ("A4", "LANDSCAPE"): QPageSize.PageSizeId.A4,
        ("A3", "PORTRAIT"): QPageSize.PageSizeId.A3,
        ("A3", "LANDSCAPE"): QPageSize.PageSizeId.A3,
    }
    for paper, orientation in expected_sizes:
        options = PdfOptions(paper, orientation)
        for kind, exporter in (("checklist", export_checklist_pdf), ("settlement", export_settlement_pdf)):
            path = exporter(db, event_id, tmp_path / f"{kind}-{paper}-{orientation}.pdf", options)
            assert path.read_bytes().startswith(b"%PDF")
            assert path.stat().st_size > 5_000
            document = QPdfDocument()
            assert document.load(str(path)) == QPdfDocument.Error.None_
            assert document.pageCount() >= 1
            points = document.pagePointSize(0)
            portrait = orientation == "PORTRAIT"
            assert (points.height() > points.width()) is portrait
            expected = QPageSize(expected_sizes[(paper, orientation)]).sizePoints()
            assert abs(min(points.width(), points.height()) - min(expected.width(), expected.height())) < 2
            assert abs(max(points.width(), points.height()) - max(expected.width(), expected.height())) < 2
            image = document.render(0, document.pagePointSize(0).toSize())
            assert not image.isNull()
            document.close()
    app.processEvents()


def test_landscape_checklist_header_has_order_and_fills_the_full_table_width():
    headers, widths = _checklist_standard_columns(789.5)
    assert headers[:3] == ["순서", "대분류", "중분류"]
    assert headers[-1] == "전화번호"
    assert "수량" in headers and "단위" in headers
    assert len(headers) == len(widths) == 14
    assert widths[0] == 24
    assert abs(sum(widths) - 789.5) < 1e-9


def test_pdf_default_filename_and_same_day_sequence(db, tmp_path):
    event_id = _event(db)
    event = db.one("SELECT * FROM events WHERE id=?", (event_id,))
    name = default_pdf_filename(event, "checklist", PdfOptions("A3", "LANDSCAPE"), date(2026, 8, 13))
    assert name == "체크리스트_내보내기 행사_20260813_A3가로.pdf"
    assert default_pdf_filename(
        event, "checklist", PdfOptions("A4", "PORTRAIT"), date(2026, 8, 13), major="시스템",
    ) == "체크리스트_내보내기 행사_시스템_20260813_A4세로.pdf"
    assert default_pdf_filename(
        event, "calendar", PdfOptions("A4", "LANDSCAPE"), date(2026, 8, 13),
        major="시스템", minor="무대",
    ) == "달력_내보내기 행사_시스템_무대_20260813_A4가로.pdf"
    assert default_pdf_filename(
        event, "checklist", PdfOptions("A4", "PORTRAIT"), date(2026, 8, 13),
        scope_label="영상 업체",
    ) == "체크리스트_내보내기 행사_영상 업체_20260813_A4세로.pdf"
    first = tmp_path / name
    first.write_bytes(b"first")
    second = next_available_pdf_path(first)
    second.write_bytes(b"second")
    assert second.name.endswith("_2.pdf")
    assert next_available_pdf_path(first).name.endswith("_3.pdf")


def test_calendar_pdf_exports_filtered_month_and_overflow_pages(db, tmp_path):
    app = QApplication.instance() or QApplication([])
    event_id = _event(db)
    for index in range(9):
        db.execute(
            """INSERT INTO event_tasks(event_id,major,minor,name,status,planned_start,due_date,sort_order)
               VALUES (?,?,?,?,?,?,?,?)""",
            (event_id, "행사", "공연", f"같은 주 겹침 일정 {index + 1}", "진행중",
             "2026-09-06", "2026-09-12", 1000 + index),
        )
    path = export_calendar_pdf(
        db, event_id, tmp_path / "calendar.pdf", 2026, 9,
        PdfOptions("A4", "LANDSCAPE"), "행사", "공연",
    )
    document = QPdfDocument()
    assert document.load(str(path)) == QPdfDocument.Error.None_
    assert document.pageCount() == 3
    for page in range(document.pageCount()):
        assert not document.render(page, document.pagePointSize(page).toSize()).isNull()
    document.close(); app.processEvents()


def test_checklist_pdf_exports_only_selected_major(db, tmp_path):
    app = QApplication.instance() or QApplication([])
    event_id = _event(db)
    db.execute("UPDATE event_tasks SET major='시스템' WHERE id=(SELECT MIN(id) FROM event_tasks WHERE event_id=?)", (event_id,))
    db.execute("UPDATE event_tasks SET major='시설' WHERE id=(SELECT MAX(id) FROM event_tasks WHERE event_id=?)", (event_id,))
    path = export_checklist_pdf(
        db, event_id, tmp_path / "checklist-system.pdf", PdfOptions("A4", "PORTRAIT"), "시스템",
    )
    document = QPdfDocument()
    assert document.load(str(path)) == QPdfDocument.Error.None_
    assert document.pageCount() == 1
    assert not document.render(0, document.pagePointSize(0).toSize()).isNull()
    document.close(); app.processEvents()


def test_checklist_excel_includes_quantity_and_unit_columns(db, tmp_path):
    """엑셀 체크리스트에도 PDF 처럼 수량·단위 열이 출력된다(가로·세로 모두)."""
    event_id = _event(db)
    db.execute(
        "UPDATE event_tasks SET quantity=3,unit='대' WHERE id=(SELECT MIN(id) FROM event_tasks WHERE event_id=?)",
        (event_id,),
    )
    for paper, orientation in (("A3", "LANDSCAPE"), ("A4", "PORTRAIT")):
        path = export_excel(
            db, tmp_path / f"checklist-qty-{paper}-{orientation}.xlsx", event_id, "checklist",
            PdfOptions(paper, orientation),
        )
        workbook = load_workbook(path, read_only=False, data_only=False)
        sheet = workbook["체크리스트"]
        headers = [sheet.cell(6, column).value for column in range(1, sheet.max_column + 1)]
        assert "수량" in headers and "단위" in headers
        # 첫 데이터 행(7행)의 수량·단위 값이 실제 업무값과 일치한다.
        qty_col = headers.index("수량") + 1
        unit_col = headers.index("단위") + 1
        assert sheet.cell(7, qty_col).value == "3"
        assert sheet.cell(7, unit_col).value == "대"
        workbook.close()
