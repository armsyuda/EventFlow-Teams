from __future__ import annotations

from datetime import date, datetime
from math import ceil
from pathlib import Path
from unicodedata import east_asian_width

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .pdf_export import PdfOptions, settlement_header_summary
from .services import EventService


INK = "212124"
MUTED = "686B70"
SUBTLE = "868B94"
LINE = "E5E7EB"
SOFT = "F7F8FA"
HEADER = "EEF0F3"
BRAND = "F25B24"
BRAND_DARK = "D84B18"
BRAND_SOFT = "FFF0E8"
MINOR_SOFT = "F2F3F5"
STATUS_COLORS = {
    "완료": ("18864B", "E8F7EF"),
    "진행중": ("1769AA", "EAF3FB"),
    "확인요청": ("9A6700", "FFF5CC"),
    "미착수": (MUTED, MINOR_SOFT),
}
THIN_BORDER = Border(
    left=Side(style="thin", color=LINE), right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE), bottom=Side(style="thin", color=LINE),
)


def _safe_filename(value: str) -> str:
    clean = "".join("_" if char in '<>:"/\\|?*' else char for char in str(value)).strip(" .")
    return clean[:80].rstrip(" ._") or "행사"


def default_excel_filename(event, kind: str, options: PdfOptions = PdfOptions(), printed_on=None,
                           major: str = "", minor: str = "") -> str:
    label = "체크리스트" if kind == "checklist" else "정산내역"
    output_date = (printed_on or date.today()).strftime("%Y%m%d")
    orientation = "세로" if options.orientation == "PORTRAIT" else "가로"
    parts = [label, _safe_filename(event["name"])]
    if major:
        parts.append(_safe_filename(major))
    if minor:
        parts.append(_safe_filename(minor))
    parts.extend([output_date, f"{options.paper}{orientation}"])
    return "_".join(parts) + ".xlsx"


def next_available_excel_path(path: Path) -> Path:
    path = Path(path)
    if not path.exists():
        return path
    index = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def _event_period(event) -> str:
    period = str(event["start_date"] or "")
    if event["end_date"]:
        period += f" - {str(event['end_date'])[5:]}"
    return period or "미입력"


def _progress(tasks) -> str:
    counts = {key: 0 for key in ("완료", "진행중", "확인요청")}
    for task in tasks:
        if task["status"] in counts:
            counts[task["status"]] += 1
    return f"완료 {counts['완료']} · 진행 {counts['진행중']} · 확인 {counts['확인요청']}"


def _money(value) -> str:
    return f"{int(value or 0):,}원"


def _apply_top_header(ws, event, document_title: str, context: str, summary: str, last_col: int) -> None:
    last = get_column_letter(last_col)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=BRAND)
    ws.row_dimensions[1].height = 5

    ws["A2"] = "EVENT FLOW"
    ws["A2"].font = Font(name="맑은 고딕", size=7, bold=True, color=BRAND)
    ws["B2"] = document_title
    ws["B2"].font = Font(name="맑은 고딕", size=7, bold=True, color=MUTED)
    timestamp_start = max(2, last_col - 2)
    ws.merge_cells(start_row=2, start_column=timestamp_start, end_row=2, end_column=last_col)
    timestamp = ws.cell(2, timestamp_start, f"출력  {datetime.now():%Y.%m.%d  %H:%M}")
    timestamp.font = Font(name="맑은 고딕", size=6, color=SUBTLE)
    timestamp.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 16

    summary_cols = 3 if last_col >= 10 else 2
    period_cols = 2 if last_col < 10 else 3
    summary_start = last_col - summary_cols + 1
    period_start = summary_start - period_cols
    title_end = max(2, period_start - 1)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=title_end)
    ws.merge_cells(start_row=3, start_column=period_start, end_row=3, end_column=summary_start - 1)
    ws.merge_cells(start_row=3, start_column=summary_start, end_row=3, end_column=last_col)
    title = ws["A3"]
    title.value = event["name"]
    title_length = len(str(event["name"] or ""))
    title_size = 13 if title_length <= 28 else (11 if title_length <= 50 else 9)
    title.font = Font(name="맑은 고딕", size=title_size, bold=True, color=INK)
    title.alignment = Alignment(vertical="center", shrink_to_fit=True)
    period = ws.cell(3, period_start, f"행사기간  {_event_period(event)}")
    period.font = Font(name="맑은 고딕", size=7, bold=True, color=MUTED)
    period.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
    summary_cell = ws.cell(3, summary_start, summary)
    summary_cell.font = Font(name="맑은 고딕", size=8, bold=True, color=BRAND_DARK)
    summary_cell.alignment = Alignment(
        horizontal="right", vertical="center", shrink_to_fit="\n" not in summary,
        wrap_text="\n" in summary,
    )
    ws.row_dimensions[3].height = 30 if "\n" in summary else 28

    ws.merge_cells(f"A4:{last}4")
    ws["A4"] = context
    ws["A4"].font = Font(name="맑은 고딕", size=7, bold=True, color=MUTED)
    ws["A4"].alignment = Alignment(vertical="center")
    ws.row_dimensions[4].height = 14
    ws.merge_cells(f"A5:{last}5")
    ws["A5"].border = Border(bottom=Side(style="thin", color=LINE))
    ws.row_dimensions[5].height = 5


def _style_table_header(ws, headers, widths) -> None:
    for column, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(6, column, header)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(name="맑은 고딕", size=7, bold=True, color=MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.row_dimensions[6].height = 30


def _configure_print(ws, options: PdfOptions, last_col: int, last_row: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.freeze_panes = "A7"
    ws.page_setup.orientation = "portrait" if options.orientation == "PORTRAIT" else "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4 if options.paper == "A4" else ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.24, right=0.24, top=0.28, bottom=0.28, header=0.08, footer=0.12)
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:6"
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.oddFooter.center.text = "Event Flow  ·  &[Page] / &[Pages]"
    ws.oddFooter.center.size = 7
    ws.oddFooter.center.color = SUBTLE


def _style_body_cell(cell, *, font_size=8, bold=False, horizontal="center", wrap=False, fill=None):
    cell.font = Font(name="맑은 고딕", size=font_size, bold=bold, color=INK)
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap, shrink_to_fit=not wrap)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _wrapped_line_count(value, column_width: float, font_size: float) -> int:
    """Estimate Excel's wrapped lines, including explicit line breaks and Korean glyph width."""
    text = str(value or "")
    capacity = max(1.0, column_width * 2.0 * (8.0 / font_size) * 0.92)
    lines = 0
    for part in text.split("\n") or [""]:
        visual_width = sum(2 if east_asian_width(char) in {"W", "F", "A"} else 1 for char in part)
        lines += max(1, ceil(visual_width / capacity))
    return lines


def _content_row_height(font_size: float, lines: int, minimum: float = 21) -> float:
    return max(minimum, round(lines * font_size * 1.45 + 4, 1))


def _merge_group_column(ws, rows, column: int, key_index: int, fill: str, font_color: str,
                        start_row: int = 7) -> None:
    start = 0
    while start < len(rows):
        value = rows[start][key_index]
        end = start + 1
        while end < len(rows) and rows[end][key_index] == value:
            end += 1
        first_row, last_row = start_row + start, start_row + end - 1
        if last_row > first_row:
            ws.merge_cells(start_row=first_row, start_column=column, end_row=last_row, end_column=column)
        cell = ws.cell(first_row, column)
        cell.value = value
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="맑은 고딕", size=8, bold=True, color=font_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        start = end


def _checklist_sheet(wb, event, tasks, options: PdfOptions, major: str, minor: str):
    ws = wb.active
    ws.title = "체크리스트"
    portrait = options.paper == "A4" and options.orientation == "PORTRAIT"
    if portrait:
        # 순서/대분류/중분류/항목/세부내용/수량/단위/진행/시작일업체/마감일업체담당자/PM담당자전화번호
        headers = ["순서", "대분류", "중분류", "항목", "세부내용", "수량", "단위", "진행",
                   "시작일\n업체", "마감일\n업체담당자", "PM 담당자\n전화번호"]
        widths = [5, 8, 10, 16, 30, 7, 7, 8, 16, 17, 17]
    else:
        headers = ["순서", "대분류", "중분류", "항목", "세부내용", "수량", "단위", "상태",
                   "작업 시작일", "마감일", "PM 담당자", "업체", "업체담당자", "전화번호"]
        widths = [5, 9, 11, 17, 34, 7, 7, 9, 11, 11, 11, 13, 13, 14]
    context = "실행 업무 현황"
    if minor:
        context += f" · {major} > {minor}"
    elif major:
        context += f" · 대분류 {major}"
    else:
        context += " · 전체"
    _apply_top_header(ws, event, "행사 체크리스트", context, _progress(tasks), len(headers))
    _style_table_header(ws, headers, widths)
    font_size = 8 if len(tasks) <= (28 if portrait else 24) else 10
    group_rows = []
    for index, task in enumerate(tasks, 1):
        row = 6 + index
        qty = str(int(task["quantity"] or 0))
        unit = task["unit"] or "식"
        if portrait:
            values = [
                index, task["major"], task["minor"], task["name"], task["detail"] or "", qty, unit, task["status"],
                f"{task['planned_start'] or '미입력'}\n{task['vendor_name'] or '미지정'}",
                f"{task['due_date'] or '미입력'}\n{task['assignee_name'] or '미지정'}",
                f"{task['pm_assignee_name'] or '미지정'}\n{task['assignee_phone'] or ''}",
            ]
        else:
            values = [
                index, task["major"], task["minor"], task["name"], task["detail"] or "", qty, unit, task["status"],
                task["planned_start"] or "미입력", task["due_date"] or "미입력",
                task["pm_assignee_name"] or "미지정", task["vendor_name"] or "미지정",
                task["assignee_name"] or "미지정", task["assignee_phone"] or "",
            ]
        group_rows.append(values)
        row_fill = None if index % 2 else SOFT
        for column, value in enumerate(values, 1):
            cell = ws.cell(row, column, value)
            # 세부내용은 좌측 정렬, 나머지는 가운데(수량·단위 포함).
            is_detail = column == 5
            _style_body_cell(
                cell, font_size=font_size, bold=column == 4,
                horizontal="left" if is_detail else "center", wrap=is_detail or (portrait and column >= 9),
                fill=row_fill,
            )
        fg, bg = STATUS_COLORS.get(task["status"], (MUTED, MINOR_SOFT))
        status = ws.cell(row, 8)
        status.fill = PatternFill("solid", fgColor=bg)
        status.font = Font(name="맑은 고딕", size=max(7, font_size - 1), bold=True, color=fg)
        detail_lines = _wrapped_line_count(task["detail"], widths[4], font_size)
        if portrait:
            stacked_lines = max(
                _wrapped_line_count(values[column - 1], widths[column - 1], font_size)
                for column in (9, 10, 11)
            )
            ws.row_dimensions[row].height = _content_row_height(
                font_size, max(2, detail_lines, stacked_lines), 28 if font_size == 8 else 34,
            )
        else:
            ws.row_dimensions[row].height = _content_row_height(font_size, detail_lines)
    _merge_group_column(ws, group_rows, 2, 1, BRAND_SOFT, BRAND_DARK)
    _merge_group_column(ws, group_rows, 3, 2, MINOR_SOFT, INK)
    last_row = max(7, 6 + len(tasks))
    _configure_print(ws, options, len(headers), last_row)
    return ws


def _settlement_sheet(wb, event, summary, options: PdfOptions):
    ws = wb.active
    ws.title = "정산내역"
    portrait = options.paper == "A4" and options.orientation == "PORTRAIT"
    if portrait:
        headers = ["대분류", "중분류", "항목", "세부내용", "수량", "단위", "단가", "공급가", "VAT", "VAT액", "합계", "업체"]
        widths = [8, 9, 18, 32, 7, 7, 13, 14, 8, 13, 14, 15]
        detail_col, qty_col, price_col, supply_col, vat_type_col, vat_col, total_col = 4, 5, 7, 8, 9, 10, 11
    else:
        headers = ["대분류", "중분류", "항목", "수량", "단위", "행사 단가", "공급가", "VAT", "VAT 금액", "합계", "업체", "세부내용"]
        widths = [9, 10, 20, 8, 8, 14, 15, 9, 14, 15, 16, 34]
        detail_col, qty_col, price_col, supply_col, vat_type_col, vat_col, total_col = 12, 4, 6, 7, 8, 9, 10
    _apply_top_header(
        ws, event, "행사 정산내역", "공급가 및 VAT 정산",
        settlement_header_summary(summary), len(headers),
    )
    _style_table_header(ws, headers, widths)
    font_size = 8 if len(summary["items"]) <= (28 if portrait else 24) else 10
    row = 7
    major_subtotal_rows = []
    current_major = None
    run_start = row
    group_values = []

    def finish_major(end_row: int):
        nonlocal row, run_start, group_values
        if current_major is None:
            return
        _merge_group_column(ws, group_values, 1, 0, BRAND_SOFT, BRAND_DARK, run_start)
        _merge_group_column(ws, group_values, 2, 1, MINOR_SOFT, INK, run_start)
        label_end = supply_col - 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_end)
        label = ws.cell(row, 1, f"{current_major} 소계")
        label.fill = PatternFill("solid", fgColor=BRAND_SOFT)
        label.font = Font(name="맑은 고딕", size=8, bold=True, color=BRAND_DARK)
        label.alignment = Alignment(horizontal="center", vertical="center")
        supply_letter = get_column_letter(supply_col)
        vat_letter = get_column_letter(vat_col)
        total_letter = get_column_letter(total_col)
        for column, letter in ((supply_col, supply_letter), (vat_col, vat_letter), (total_col, total_letter)):
            cell = ws.cell(row, column, f"=SUM({letter}{run_start}:{letter}{end_row})")
            cell.fill = PatternFill("solid", fgColor=BRAND_SOFT)
            cell.font = Font(name="맑은 고딕", size=8, bold=True, color=BRAND_DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "#,##0"
        for column in range(1, len(headers) + 1):
            ws.cell(row, column).border = THIN_BORDER
            if ws.cell(row, column).fill.fill_type is None:
                ws.cell(row, column).fill = PatternFill("solid", fgColor=BRAND_SOFT)
        major_subtotal_rows.append(row)
        ws.row_dimensions[row].height = 22
        row += 1
        run_start = row
        group_values = []

    for item in summary["items"]:
        if current_major is not None and item["major"] != current_major:
            finish_major(row - 1)
        current_major = item["major"]
        vat_label = "과세" if item["vat_type"] == "TAXABLE" else "면세"
        detail = item["detail"] or item["note"] or ""
        if portrait:
            values = [item["major"], item["minor"], item["name"], detail, item["quantity"] or 0,
                      item["unit"] or "식", item["unit_price"] or 0, None, vat_label, None, None,
                      item["vendor_name"] or "미지정"]
        else:
            values = [item["major"], item["minor"], item["name"], item["quantity"] or 0,
                      item["unit"] or "식", item["unit_price"] or 0, None, vat_label, None, None,
                      item["vendor_name"] or "미지정", detail]
        group_values.append(values)
        row_fill = None if (row - 7) % 2 == 0 else SOFT
        for column, value in enumerate(values, 1):
            cell = ws.cell(row, column, value)
            is_detail = column == detail_col
            _style_body_cell(cell, font_size=font_size, bold=column == 3,
                             horizontal="left" if is_detail else "center", wrap=is_detail, fill=row_fill)
        qty_letter = get_column_letter(qty_col)
        price_letter = get_column_letter(price_col)
        supply_letter = get_column_letter(supply_col)
        vat_type_letter = get_column_letter(vat_type_col)
        vat_letter = get_column_letter(vat_col)
        ws.cell(row, supply_col, f"=ROUND({qty_letter}{row}*{price_letter}{row},0)")
        ws.cell(row, vat_col, f'=IF({vat_type_letter}{row}="과세",ROUND({supply_letter}{row}*0.1,0),0)')
        ws.cell(row, total_col, f"={supply_letter}{row}+{vat_letter}{row}")
        for column in (qty_col, price_col, supply_col, vat_col, total_col):
            ws.cell(row, column).number_format = "#,##0"
        detail_lines = _wrapped_line_count(detail, widths[detail_col - 1], font_size)
        ws.row_dimensions[row].height = _content_row_height(font_size, detail_lines)
        row += 1
    if current_major is not None:
        finish_major(row - 1)

    total_row = row
    label_end = supply_col - 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=label_end)
    ws.cell(total_row, 1, "전체 합계")
    subtotal_refs = {
        supply_col: [f"{get_column_letter(supply_col)}{value}" for value in major_subtotal_rows],
        vat_col: [f"{get_column_letter(vat_col)}{value}" for value in major_subtotal_rows],
        total_col: [f"{get_column_letter(total_col)}{value}" for value in major_subtotal_rows],
    }
    for column in range(1, len(headers) + 1):
        cell = ws.cell(total_row, column)
        cell.fill = PatternFill("solid", fgColor=BRAND_SOFT)
        cell.border = THIN_BORDER
        cell.font = Font(name="맑은 고딕", size=9, bold=True, color=BRAND_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column, refs in subtotal_refs.items():
        ws.cell(total_row, column, f"=SUM({','.join(refs)})" if refs else "=0")
        ws.cell(total_row, column).number_format = "#,##0"
    ws.row_dimensions[total_row].height = 25
    _configure_print(ws, options, len(headers), total_row)
    return ws


def export_excel(db, destination: Path, event_id: int, kind: str = "checklist",
                 options: PdfOptions = PdfOptions(), major: str = "", minor: str = ""):
    if kind not in {"checklist", "settlement"}:
        raise ValueError("Excel 내보내기는 체크리스트 또는 정산내역만 지원합니다.")
    service = EventService(db)
    event = service.get_event(event_id)
    if not event:
        raise ValueError("내보낼 행사를 찾을 수 없습니다.")
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    if kind == "checklist":
        tasks = [dict(row) for row in service.list_tasks(event_id, major=major)]
        if minor:
            tasks = [task for task in tasks if task["minor"] == minor]
        _checklist_sheet(wb, event, tasks, options, major, minor)
    else:
        _settlement_sheet(wb, event, service.settlement_summary(event_id), options)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(destination)
    check = load_workbook(destination, read_only=False, data_only=False)
    if check.sheetnames != (["체크리스트"] if kind == "checklist" else ["정산내역"]):
        raise OSError("Excel 시트 구성이 올바르지 않습니다.")
    check.close()
    return destination
